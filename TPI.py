from datetime import datetime


def parsear_fecha(fecha_texto):
    """Convierte una cadena DD/MM/AAAA a un objeto datetime.date."""
    try:
        return datetime.strptime(fecha_texto, "%d/%m/%Y").date()
    except ValueError:
        raise ValueError("Formato incorrecto. Use DD/MM/AAAA.")


def buscar_empleado(hoja_empleados, id_buscado):
    """Busca un empleado por ID en la hoja Empleados."""
    for fila_idx, fila in enumerate(hoja_empleados.iter_rows(min_row=2, values_only=True), start=2):
        if str(fila[0]) == str(id_buscado):
            return {
                "row": fila_idx,
                "id": fila[0],
                "nombre": fila[1],
                "saldo": fila[2],
            }
    return None


def registrar_solicitud(hoja_solicitudes, solicitud):
    """Agrega un registro de solicitud en la hoja Solicitudes."""
    hoja_solicitudes.append(solicitud)


def confirmar_accion(mensaje):
    """Pregunta al usuario si desea continuar con una acción."""
    respuesta = input(mensaje).strip().lower()
    return respuesta in ("s", "si")


def cargar_hojas_excel(ruta_archivo):
    """Carga el libro y devuelve las hojas necesarias."""
    from openpyxl import load_workbook

    libro = load_workbook(ruta_archivo)
    hoja_empleados = libro["Empleados"]

    if "Solicitudes" in libro.sheetnames:
        hoja_solicitudes = libro["Solicitudes"]
    else:
        hoja_solicitudes = libro.create_sheet("Solicitudes")
        hoja_solicitudes.append([
            "id",
            "nombre",
            "fecha_desde",
            "fecha_hasta",
            "dias",
            "estado",
            "fecha_solicitud",
        ])

    return libro, hoja_empleados, hoja_solicitudes


def main():
    print("\n=== Sistema de Vacaciones ===")

    estado = "INICIO"
    datos = {
        "empleado_id": None,
        "fecha_desde_texto": None,
        "fecha_hasta_texto": None,
        "fecha_desde": None,
        "fecha_hasta": None,
        "dias_solicitados": None,
        "empleado": None,
        "estado_result": None,
        "hoja_empleados": None,
        "hoja_solicitudes": None,
    }

    while True:
        if estado == "INICIO":
            estado = "ESPERANDO_FECHAS"

        elif estado == "ESPERANDO_FECHAS":
            datos["empleado_id"] = input("Ingrese su ID: ").strip()
            datos["fecha_desde_texto"] = input("Fecha desde (DD/MM/AAAA): ").strip()
            datos["fecha_hasta_texto"] = input("Fecha hasta (DD/MM/AAAA): ").strip()
            estado = "VALIDANDO_FORMATO"

        elif estado == "VALIDANDO_FORMATO":
            try:
                datos["fecha_desde"] = parsear_fecha(datos["fecha_desde_texto"])
                datos["fecha_hasta"] = parsear_fecha(datos["fecha_hasta_texto"])
            except ValueError as error:
                print(error)
                print("Por favor, reingrese las fechas.")
                estado = "ESPERANDO_FECHAS"
                continue

            if datos["fecha_hasta"] < datos["fecha_desde"]:
                print("La fecha final no puede ser anterior a la fecha inicial.")
                estado = "ESPERANDO_FECHAS"
                continue

            hoy = datetime.now().date()
            if datos["fecha_desde"] < hoy:
                print("No se permiten fechas retroactivas.")
                estado = "ESPERANDO_FECHAS"
                continue

            datos["dias_solicitados"] = (datos["fecha_hasta"] - datos["fecha_desde"]).days + 1
            print(f"Se solicitaron {datos['dias_solicitados']} días.")
            estado = "CONSULTANDO_SALDO"

        elif estado == "CONSULTANDO_SALDO":
            try:
                libro, datos["hoja_empleados"], datos["hoja_solicitudes"] = cargar_hojas_excel("base_vacaciones.xlsx")
            except Exception as error:
                print("No se pudo abrir 'base_vacaciones.xlsx':", error)
                return

            datos["empleado"] = buscar_empleado(datos["hoja_empleados"], datos["empleado_id"])
            if datos["empleado"] is None:
                print("Empleado inexistente")
                datos["estado_result"] = "RECHAZADO - EMPLEADO INEXISTENTE"
                registrar_solicitud(
                    datos["hoja_solicitudes"],
                    [
                        datos["empleado_id"],
                        None,
                        datos["fecha_desde_texto"],
                        datos["fecha_hasta_texto"],
                        datos["dias_solicitados"],
                        datos["estado_result"],
                        datetime.now().strftime("%d/%m/%Y %H:%M"),
                    ],
                )
                libro.save("base_vacaciones.xlsx")
                return

            estado = "CALCULANDO_DIAS"

        elif estado == "CALCULANDO_DIAS":
            estado = "EVALUANDO_SALDO"

        elif estado == "EVALUANDO_SALDO":
            empleado = datos["empleado"]
            saldo = empleado["saldo"]
            nombre = empleado["nombre"]
            print(f"Empleado: {nombre}")
            print(f"Saldo disponible: {saldo} días")

            if datos["dias_solicitados"] <= saldo:
                if confirmar_accion("Confirmar solicitud y descontar días? (s/n): "):
                    datos["estado_result"] = "APROBADO"
                    datos["hoja_empleados"].cell(
                        row=empleado["row"], column=3, value=saldo - datos["dias_solicitados"]
                    )
                    print(f"Solicitud aprobada para {nombre}.")
                else:
                    datos["estado_result"] = "CANCELADO"
                    print(f"Solicitud cancelada para {nombre}.")
            else:
                datos["estado_result"] = "RECHAZADO"
                print(f"Saldo insuficiente para {nombre}.")

            registrar_solicitud(
                datos["hoja_solicitudes"],
                [
                    datos["empleado_id"],
                    nombre,
                    datos["fecha_desde_texto"],
                    datos["fecha_hasta_texto"],
                    datos["dias_solicitados"],
                    datos["estado_result"],
                    datetime.now().strftime("%d/%m/%Y %H:%M"),
                ],
            )
            libro.save("base_vacaciones.xlsx")
            return

        else:
            print(f"Estado desconocido: {estado}")
            return


if __name__ == "__main__":
    main()